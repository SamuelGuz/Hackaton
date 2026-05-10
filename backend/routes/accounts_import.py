"""Accounts import endpoints (JSON + CSV/XLSX) for accounts and related datasets."""

from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError

from backend.routes.schemas_import import (
    ImportConversationRow,
    ImportConversationsRequest,
    ImportError,
    ImportTicketRow,
    ImportTicketsRequest,
    ImportRequest,
    ImportResponse,
    ImportUsageEventRow,
    ImportUsageEventsRequest,
    RelatedImportResponse,
)
from backend.shared.api_auth import require_api_key
from backend.shared.supabase_client import get_client

router = APIRouter(prefix="/accounts", tags=["accounts"])

_IMPORT_VERSION = "excel-import-v1"


def _http_error(status: int, code: str, message: str, details: dict[str, Any] | None = None) -> HTTPException:
    return HTTPException(
        status_code=status,
        detail={"error": code, "message": message, "details": details or {}},
    )


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace(" ", "_").replace("-", "_")


def _parse_upload_rows(file: UploadFile) -> list[dict[str, Any]]:
    content = file.file.read()
    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        out: list[dict[str, Any]] = []
        for row in reader:
            parsed: dict[str, Any] = {}
            for key, val in row.items():
                if key is None:
                    continue
                parsed[_normalize_header(key)] = val.strip() if isinstance(val, str) else val
            out.append(parsed)
        return out
    if filename.endswith(".xlsx"):
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_normalize_header(str(h or "")) for h in rows[0]]
        out: list[dict[str, Any]] = []
        for values in rows[1:]:
            item: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = values[idx] if idx < len(values) else None
                item[header] = value
            out.append(item)
        return out
    raise _http_error(400, "unsupported_file_type", "Solo se soportan archivos .csv o .xlsx")


def _create_health_records(
    *,
    client: Any,
    account_id: str,
    churn_risk_score: int,
    expansion_score: int,
    health_status: str,
    predicted_churn_reason: str | None,
    crystal_ball_reasoning: str,
) -> None:
    now_iso = datetime.now(timezone.utc).isoformat()
    snapshot_payload = {
        "account_id": account_id,
        "churn_risk_score": churn_risk_score,
        "expansion_score": expansion_score,
        "health_status": health_status,
        "top_signals": [],
        "predicted_churn_reason": predicted_churn_reason,
        "crystal_ball_confidence": None,
        "crystal_ball_reasoning": crystal_ball_reasoning,
        "ready_to_expand": expansion_score >= 60,
        "recommended_plan": None,
        "expansion_reasoning": None,
        "suggested_upsell_message": None,
        "computed_at": now_iso,
        "computed_by_version": _IMPORT_VERSION,
    }
    history_payload = {
        "id": str(uuid.uuid4()),
        "account_id": account_id,
        "churn_risk_score": churn_risk_score,
        "expansion_score": expansion_score,
        "health_status": health_status,
        "top_signals": [],
        "predicted_churn_reason": predicted_churn_reason,
        "crystal_ball_confidence": None,
        "computed_at": now_iso,
        "computed_by_version": _IMPORT_VERSION,
    }
    client.table("account_health_snapshot").upsert(snapshot_payload, on_conflict="account_id").execute()
    client.table("account_health_history").insert(history_payload).execute()


def _import_accounts_rows(rows: list[Any]) -> ImportResponse:
    client = get_client()

    csm_rows = client.table("csm_team").select("id,name").execute().data or []
    csm_map: dict[str, str] = {
        str(row["name"]).strip().lower(): str(row["id"])
        for row in csm_rows
        if row.get("name") and row.get("id")
    }
    csm_ids = {str(row["id"]) for row in csm_rows if row.get("id")}

    existing_rows = client.table("accounts").select("id,account_number").execute().data or []
    existing_numbers: set[str] = {
        str(row.get("account_number") or "").strip() for row in existing_rows if row.get("account_number")
    }
    inserted_ids: list[str] = []
    errors: list[ImportError] = []
    skipped = 0

    for idx, row in enumerate(rows):
        number = row.account_number.strip()
        if number in existing_numbers:
            skipped += 1
            continue

        csm_id = row.csm_id.strip() if row.csm_id else None
        if csm_id and csm_id not in csm_ids:
            errors.append(ImportError(row_index=idx, key=number, message=f"csm_id inválido: {csm_id}"))
            continue
        if not csm_id:
            csm_lookup = (row.csm_assigned or "").strip().lower()
            csm_id = csm_map.get(csm_lookup)
        if not csm_id:
            errors.append(
                ImportError(
                    row_index=idx,
                    key=number,
                    message=f"CSM '{row.csm_assigned}' no encontrado",
                )
            )
            continue

        account_id = str(uuid.uuid4())
        account_payload = {
            "id": account_id,
            "account_number": number,
            "name": row.name.strip(),
            "industry": row.industry,
            "size": row.size,
            "geography": row.geography,
            "plan": row.plan,
            "arr_usd": row.arr_usd,
            "seats_purchased": row.seats_purchased,
            "seats_active": row.seats_active,
            "signup_date": row.signup_date.isoformat(),
            "contract_renewal_date": row.contract_renewal_date.isoformat(),
            "champion_name": row.champion_name,
            "champion_email": row.champion_email.strip(),
            "champion_role": row.champion_role,
            "champion_phone": row.champion_phone.strip() if row.champion_phone else None,
            "champion_changed_recently": row.champion_changed_recently,
            "csm_id": csm_id,
            "last_qbr_date": row.last_qbr_date.isoformat() if row.last_qbr_date else None,
            "current_nps_score": row.current_nps_score,
            "current_nps_category": row.current_nps_category,
            "last_nps_at": row.last_nps_at.isoformat() if row.last_nps_at else None,
        }

        try:
            client.table("accounts").insert(account_payload).execute()
        except Exception as exc:
            errors.append(
                ImportError(
                    row_index=idx,
                    key=number,
                    message=f"insert account falló: {exc}",
                )
            )
            continue

        inserted_ids.append(account_id)
        existing_numbers.add(number)
        try:
            _create_health_records(
                client=client,
                account_id=account_id,
                churn_risk_score=row.churn_risk_score if row.churn_risk_score is not None else 25,
                expansion_score=row.expansion_score if row.expansion_score is not None else 30,
                health_status=row.health_status or "stable",
                predicted_churn_reason=row.predicted_churn_reason,
                crystal_ball_reasoning=(
                    row.crystal_ball_reasoning or "Importado por API; sin análisis Crystal Ball."
                ),
            )
        except Exception as exc:
            errors.append(
                ImportError(
                    row_index=idx,
                    key=number,
                    message=f"snapshot/history falló: {exc}",
                )
            )

    return ImportResponse(
        inserted=len(inserted_ids),
        skipped=skipped,
        errors=errors,
        inserted_ids=inserted_ids,
    )


def _resolve_account_id(
    *,
    client: Any,
    account_id: str | None,
    account_number: str | None,
    id_cache: dict[str, bool],
    number_cache: dict[str, str],
) -> tuple[str | None, str | None]:
    if account_id:
        account_id = account_id.strip()
    if account_number:
        account_number = account_number.strip()
    if account_id:
        if account_id not in id_cache:
            exists = client.table("accounts").select("id").eq("id", account_id).limit(1).execute()
            id_cache[account_id] = bool(exists.data)
        if not id_cache[account_id]:
            return None, "account_id no existe"
        if account_number:
            if account_number not in number_cache:
                by_num = (
                    client.table("accounts")
                    .select("id")
                    .eq("account_number", account_number)
                    .limit(1)
                    .execute()
                )
                if by_num.data:
                    number_cache[account_number] = str(by_num.data[0]["id"])
            linked = number_cache.get(account_number)
            if linked and linked != account_id:
                return None, "account_id y account_number no apuntan a la misma cuenta"
        return account_id, None
    if not account_number:
        return None, "Se requiere account_id o account_number"
    if account_number not in number_cache:
        by_num = (
            client.table("accounts")
            .select("id")
            .eq("account_number", account_number)
            .limit(1)
            .execute()
        )
        if by_num.data:
            number_cache[account_number] = str(by_num.data[0]["id"])
    resolved = number_cache.get(account_number)
    if not resolved:
        return None, "account_number no existe"
    return resolved, None


def _usage_duplicate_exists(client: Any, account_id: str, row: ImportUsageEventRow) -> bool:
    res = (
        client.table("usage_events")
        .select("id,feature_name,user_email")
        .eq("account_id", account_id)
        .eq("event_type", row.event_type)
        .eq("occurred_at", row.occurred_at.isoformat())
        .limit(10)
        .execute()
    )
    target_feature = row.feature_name or ""
    target_user = row.user_email or ""
    for item in res.data or []:
        if (item.get("feature_name") or "") == target_feature and (item.get("user_email") or "") == target_user:
            return True
    return False


def _ticket_duplicate_exists(client: Any, account_id: str, row: ImportTicketRow) -> bool:
    res = (
        client.table("tickets")
        .select("id")
        .eq("account_id", account_id)
        .eq("subject", row.subject)
        .eq("status", row.status)
        .eq("opened_at", row.opened_at.isoformat())
        .limit(1)
        .execute()
    )
    return bool(res.data)


def _conversation_duplicate_exists(client: Any, account_id: str, row: ImportConversationRow) -> bool:
    res = (
        client.table("conversations")
        .select("id")
        .eq("account_id", account_id)
        .eq("channel", row.channel)
        .eq("direction", row.direction)
        .eq("occurred_at", row.occurred_at.isoformat())
        .eq("content", row.content)
        .limit(1)
        .execute()
    )
    return bool(res.data)


def _import_usage_rows(rows: list[ImportUsageEventRow]) -> RelatedImportResponse:
    client = get_client()
    id_cache: dict[str, bool] = {}
    number_cache: dict[str, str] = {}
    errors: list[ImportError] = []
    inserted = 0

    for idx, row in enumerate(rows):
        resolved_id, err = _resolve_account_id(
            client=client,
            account_id=row.account_id,
            account_number=row.account_number,
            id_cache=id_cache,
            number_cache=number_cache,
        )
        key = row.account_number or row.account_id or f"row_{idx}"
        if err or not resolved_id:
            errors.append(ImportError(row_index=idx, key=key, message=err or "Cuenta no encontrada"))
            continue
        if _usage_duplicate_exists(client, resolved_id, row):
            errors.append(
                ImportError(
                    row_index=idx,
                    key=key,
                    message="Duplicado en usage_events por llave compuesta natural",
                )
            )
            continue
        payload = {
            "account_id": resolved_id,
            "event_type": row.event_type,
            "feature_name": row.feature_name,
            "user_email": row.user_email,
            "occurred_at": row.occurred_at.isoformat(),
            "metadata": row.metadata,
        }
        try:
            client.table("usage_events").insert(payload).execute()
            inserted += 1
        except Exception as exc:
            errors.append(ImportError(row_index=idx, key=key, message=f"insert usage_events falló: {exc}"))
    return RelatedImportResponse(inserted=inserted, errors=errors)


def _import_tickets_rows(rows: list[ImportTicketRow]) -> RelatedImportResponse:
    client = get_client()
    id_cache: dict[str, bool] = {}
    number_cache: dict[str, str] = {}
    errors: list[ImportError] = []
    inserted = 0

    for idx, row in enumerate(rows):
        resolved_id, err = _resolve_account_id(
            client=client,
            account_id=row.account_id,
            account_number=row.account_number,
            id_cache=id_cache,
            number_cache=number_cache,
        )
        key = row.account_number or row.account_id or f"row_{idx}"
        if err or not resolved_id:
            errors.append(ImportError(row_index=idx, key=key, message=err or "Cuenta no encontrada"))
            continue
        if _ticket_duplicate_exists(client, resolved_id, row):
            errors.append(
                ImportError(row_index=idx, key=key, message="Duplicado en tickets por llave compuesta natural")
            )
            continue
        payload = {
            "account_id": resolved_id,
            "subject": row.subject,
            "description": row.description,
            "priority": row.priority,
            "status": row.status,
            "sentiment": row.sentiment,
            "opened_at": row.opened_at.isoformat(),
            "resolved_at": row.resolved_at.isoformat() if row.resolved_at else None,
            "first_response_hours": row.first_response_hours,
        }
        try:
            client.table("tickets").insert(payload).execute()
            inserted += 1
        except Exception as exc:
            errors.append(ImportError(row_index=idx, key=key, message=f"insert tickets falló: {exc}"))
    return RelatedImportResponse(inserted=inserted, errors=errors)


def _import_conversations_rows(rows: list[ImportConversationRow]) -> RelatedImportResponse:
    client = get_client()
    id_cache: dict[str, bool] = {}
    number_cache: dict[str, str] = {}
    errors: list[ImportError] = []
    inserted = 0

    for idx, row in enumerate(rows):
        resolved_id, err = _resolve_account_id(
            client=client,
            account_id=row.account_id,
            account_number=row.account_number,
            id_cache=id_cache,
            number_cache=number_cache,
        )
        key = row.account_number or row.account_id or f"row_{idx}"
        if err or not resolved_id:
            errors.append(ImportError(row_index=idx, key=key, message=err or "Cuenta no encontrada"))
            continue
        if _conversation_duplicate_exists(client, resolved_id, row):
            errors.append(
                ImportError(
                    row_index=idx,
                    key=key,
                    message="Duplicado en conversations por llave compuesta natural",
                )
            )
            continue
        payload = {
            "account_id": resolved_id,
            "channel": row.channel,
            "direction": row.direction,
            "participants": row.participants,
            "subject": row.subject,
            "content": row.content,
            "sentiment": row.sentiment,
            "occurred_at": row.occurred_at.isoformat(),
        }
        try:
            client.table("conversations").insert(payload).execute()
            inserted += 1
        except Exception as exc:
            errors.append(ImportError(row_index=idx, key=key, message=f"insert conversations falló: {exc}"))
    return RelatedImportResponse(inserted=inserted, errors=errors)


def _parse_usage_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        raw_metadata = item.get("metadata")
        if isinstance(raw_metadata, str) and raw_metadata.strip():
            try:
                item["metadata"] = json.loads(raw_metadata)
            except json.JSONDecodeError:
                pass
        out.append(item)
    return out


def _parse_conversation_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        participants = item.get("participants")
        if isinstance(participants, str):
            sep = ";" if ";" in participants else ","
            item["participants"] = [p.strip() for p in participants.split(sep) if p.strip()]
        out.append(item)
    return out


@router.post("/import", response_model=ImportResponse)
def import_accounts(request: ImportRequest, _auth: None = Depends(require_api_key)) -> ImportResponse:
    return _import_accounts_rows(request.accounts)


@router.post("/import/file", response_model=ImportResponse)
def import_accounts_file(file: UploadFile, _auth: None = Depends(require_api_key)) -> ImportResponse:
    try:
        rows = _parse_upload_rows(file)
        request = ImportRequest(accounts=rows)
    except ValidationError as exc:
        raise _http_error(422, "validation_error", "Archivo inválido para accounts", {"errors": exc.errors()}) from exc
    return _import_accounts_rows(request.accounts)


@router.post("/import/usage-events", response_model=RelatedImportResponse)
def import_usage_events(
    request: ImportUsageEventsRequest,
    _auth: None = Depends(require_api_key),
) -> RelatedImportResponse:
    return _import_usage_rows(request.rows)


@router.post("/import/usage-events/file", response_model=RelatedImportResponse)
def import_usage_events_file(
    file: UploadFile,
    _auth: None = Depends(require_api_key),
) -> RelatedImportResponse:
    try:
        rows = _parse_usage_rows(_parse_upload_rows(file))
        request = ImportUsageEventsRequest(rows=rows)
    except ValidationError as exc:
        raise _http_error(
            422,
            "validation_error",
            "Archivo inválido para usage_events",
            {"errors": exc.errors()},
        ) from exc
    return _import_usage_rows(request.rows)


@router.post("/import/tickets", response_model=RelatedImportResponse)
def import_tickets(
    request: ImportTicketsRequest,
    _auth: None = Depends(require_api_key),
) -> RelatedImportResponse:
    return _import_tickets_rows(request.rows)


@router.post("/import/tickets/file", response_model=RelatedImportResponse)
def import_tickets_file(
    file: UploadFile,
    _auth: None = Depends(require_api_key),
) -> RelatedImportResponse:
    try:
        rows = _parse_upload_rows(file)
        request = ImportTicketsRequest(rows=rows)
    except ValidationError as exc:
        raise _http_error(422, "validation_error", "Archivo inválido para tickets", {"errors": exc.errors()}) from exc
    return _import_tickets_rows(request.rows)


@router.post("/import/conversations", response_model=RelatedImportResponse)
def import_conversations(
    request: ImportConversationsRequest,
    _auth: None = Depends(require_api_key),
) -> RelatedImportResponse:
    return _import_conversations_rows(request.rows)


@router.post("/import/conversations/file", response_model=RelatedImportResponse)
def import_conversations_file(
    file: UploadFile,
    _auth: None = Depends(require_api_key),
) -> RelatedImportResponse:
    try:
        rows = _parse_conversation_rows(_parse_upload_rows(file))
        request = ImportConversationsRequest(rows=rows)
    except ValidationError as exc:
        raise _http_error(
            422,
            "validation_error",
            "Archivo inválido para conversations",
            {"errors": exc.errors()},
        ) from exc
    return _import_conversations_rows(request.rows)
